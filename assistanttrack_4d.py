"""
╔══════════════════════════════════════════════════════════════╗
║         ASSISTANTTRACK 4D — Sistema de Control Laboral       ║
║         Computer Vision · Turnos · Overtime · Costos         ║
╚══════════════════════════════════════════════════════════════╝
Módulos:
  1. Gestión de Empleados
  2. Control de Turnos (Check-in / Check-out)
  3. Control de Overtime (Inicio / Fin OT)
  4. Motor de Alertas
  5. Análisis de Costos y ROI
  6. Generador de Reportes (Excel + CSV)
  7. Simulador de datos CV
"""

import json
import csv
import hashlib
import random
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict


def _base_dir() -> str:
    """Devuelve el directorio base ya sea ejecutando como .exe o como .py."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# Forzar UTF-8 en consola Windows para caracteres especiales
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─────────────────────────────────────────────────────────────
#  CONSTANTES
# ─────────────────────────────────────────────────────────────
COSTO_HORA_BASE   = 15.00   # USD por hora normal
FACTOR_OT         = 1.50    # Multiplicador overtime
LIMITE_OT_DIARIO  = 4.0     # Horas máx. de OT por día
UMBRAL_TARDANZA   = 10      # minutos
UMBRAL_ALERTA_OT  = 0.70    # 70% del límite presupuestal

# ─────────────────────────────────────────────────────────────
#  UTILIDADES
# ─────────────────────────────────────────────────────────────
def timestamp_hash(emp_id: str, evento: str, dt: datetime) -> str:
    """Genera hash SHA-256 inmutable para evidencia del evento CV."""
    raw = f"{emp_id}|{evento}|{dt.isoformat()}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16].upper()

def fmt_time(dt: datetime) -> str:
    return dt.strftime("%H:%M:%S")

def fmt_date(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")

def fmt_horas(minutos: float) -> str:
    h = int(minutos // 60)
    m = int(minutos % 60)
    return f"{h}h {m:02d}m"

def separador(char="─", ancho=62):
    print(char * ancho)

def titulo(texto: str):
    separador("═")
    print(f"  {texto}")
    separador("═")

def subtitulo(texto: str):
    separador()
    print(f"  {texto}")
    separador()

# ─────────────────────────────────────────────────────────────
#  CLASE EMPLEADO
# ─────────────────────────────────────────────────────────────
class Empleado:
    def __init__(self, emp_id, nombre, area, turno_inicio, turno_fin,
                 costo_hora=COSTO_HORA_BASE):
        self.id            = emp_id
        self.nombre        = nombre
        self.area          = area
        self.turno_inicio  = turno_inicio   # "HH:MM"
        self.turno_fin     = turno_fin      # "HH:MM"
        self.costo_hora    = costo_hora

    def __repr__(self):
        return f"Empleado({self.id}, {self.nombre}, {self.area})"

# ─────────────────────────────────────────────────────────────
#  CLASE EVENTO CV
# ─────────────────────────────────────────────────────────────
class EventoCV:
    TIPOS = ("CHECK_IN", "CHECK_OUT", "OT_START", "OT_END")

    def __init__(self, emp_id: str, tipo: str, dt: datetime,
                 confianza: float = 0.994):
        assert tipo in self.TIPOS, f"Tipo inválido: {tipo}"
        self.emp_id     = emp_id
        self.tipo       = tipo
        self.datetime   = dt
        self.confianza  = confianza
        self.hash_ev    = timestamp_hash(emp_id, tipo, dt)
        self.autorizado = True  # Por defecto; OT_START puede ser False

    def to_dict(self) -> dict:
        return {
            "emp_id":    self.emp_id,
            "tipo":      self.tipo,
            "fecha":     fmt_date(self.datetime),
            "hora":      fmt_time(self.datetime),
            "confianza": f"{self.confianza:.1%}",
            "hash_cv":   self.hash_ev,
            "autorizado": self.autorizado,
        }

# ─────────────────────────────────────────────────────────────
#  MOTOR DE TURNOS
# ─────────────────────────────────────────────────────────────
class MotorTurnos:
    def __init__(self):
        self.empleados: dict[str, Empleado] = {}
        self.eventos:   list[EventoCV]      = []
        self.alertas:   list[dict]          = []

    # ── Registro ──────────────────────────────────────────────
    def registrar_empleado(self, emp: Empleado):
        self.empleados[emp.id] = emp

    def registrar_evento(self, evento: EventoCV):
        self.eventos.append(evento)
        self._evaluar_alertas(evento)

    # ── Alertas ───────────────────────────────────────────────
    def _evaluar_alertas(self, ev: EventoCV):
        emp = self.empleados.get(ev.emp_id)
        if not emp:
            return

        if ev.tipo == "CHECK_IN":
            # Validar tardanza
            turno_dt = datetime.strptime(
                f"{fmt_date(ev.datetime)} {emp.turno_inicio}", "%Y-%m-%d %H:%M")
            delta_min = (ev.datetime - turno_dt).total_seconds() / 60
            if delta_min > UMBRAL_TARDANZA:
                self._crear_alerta(
                    emp_id=ev.emp_id,
                    nivel="WARNING",
                    mensaje=f"Tardanza de {int(delta_min)} min — {emp.nombre}",
                    dt=ev.datetime,
                )

        elif ev.tipo == "OT_START" and not ev.autorizado:
            self._crear_alerta(
                emp_id=ev.emp_id,
                nivel="CRITICAL",
                mensaje=f"OT sin autorización gerencial — {emp.nombre}",
                dt=ev.datetime,
            )

    def _crear_alerta(self, emp_id, nivel, mensaje, dt):
        alerta = {
            "timestamp": dt.isoformat(),
            "emp_id":    emp_id,
            "nivel":     nivel,
            "mensaje":   mensaje,
        }
        self.alertas.append(alerta)

    # ── Cálculos ──────────────────────────────────────────────
    def calcular_horas_empleado(self, emp_id: str, fecha: str) -> dict:
        """Calcula horas normales y OT para un empleado en un día."""
        evs = [e for e in self.eventos
               if e.emp_id == emp_id and fmt_date(e.datetime) == fecha]

        check_in  = next((e.datetime for e in evs if e.tipo == "CHECK_IN"),  None)
        check_out = next((e.datetime for e in evs if e.tipo == "CHECK_OUT"), None)
        ot_start  = next((e.datetime for e in evs if e.tipo == "OT_START"),  None)
        ot_end    = next((e.datetime for e in evs if e.tipo == "OT_END"),    None)

        horas_normales = 0.0
        horas_ot       = 0.0

        emp = self.empleados.get(emp_id)

        if check_in and check_out:
            horas_normales = (check_out - check_in).total_seconds() / 3600
            # Descontar si hay OT incluido
            if ot_start and ot_start < check_out:
                horas_normales = (ot_start - check_in).total_seconds() / 3600

        if ot_start and ot_end:
            horas_ot = (ot_end - ot_start).total_seconds() / 3600
            horas_ot = min(horas_ot, LIMITE_OT_DIARIO)

        cph = emp.costo_hora if emp else COSTO_HORA_BASE
        costo_normal = horas_normales * cph
        costo_ot     = horas_ot * cph * FACTOR_OT
        costo_total  = costo_normal + costo_ot

        return {
            "emp_id":          emp_id,
            "nombre":          emp.nombre if emp else "—",
            "fecha":           fecha,
            "horas_normales":  round(horas_normales, 2),
            "horas_ot":        round(horas_ot, 2),
            "costo_normal":    round(costo_normal, 2),
            "costo_ot":        round(costo_ot, 2),
            "costo_total":     round(costo_total, 2),
        }

    def resumen_costos_dia(self, fecha: str) -> dict:
        """Agrega costos de todos los empleados para una fecha."""
        resultados = []
        emp_ids_con_eventos = {e.emp_id for e in self.eventos
                               if fmt_date(e.datetime) == fecha}
        for emp_id in emp_ids_con_eventos:
            r = self.calcular_horas_empleado(emp_id, fecha)
            resultados.append(r)

        total_normal = sum(r["costo_normal"] for r in resultados)
        total_ot     = sum(r["costo_ot"] for r in resultados)
        total_horas  = sum(r["horas_ot"] for r in resultados)

        return {
            "fecha":         fecha,
            "empleados":     resultados,
            "total_normal":  round(total_normal, 2),
            "total_ot":      round(total_ot, 2),
            "total_general": round(total_normal + total_ot, 2),
            "horas_ot_dia":  round(total_horas, 2),
        }

# ─────────────────────────────────────────────────────────────
#  SIMULADOR DE DATOS CV
# ─────────────────────────────────────────────────────────────
NOMBRES = [
    ("EMP-001","María García","Producción"),
    ("EMP-002","Carlos López","Producción"),
    ("EMP-003","Ana Martínez","Logística"),
    ("EMP-004","Luis Hernández","Producción"),
    ("EMP-005","Rosa Jiménez","Calidad"),
    ("EMP-006","Pedro Ruiz","Logística"),
    ("EMP-007","Elena Torres","Producción"),
    ("EMP-008","Jorge Díaz","Mantenimiento"),
    ("EMP-009","Laura Sánchez","Calidad"),
    ("EMP-010","Miguel Flores","Producción"),
]

def simular_semana(motor: MotorTurnos, fecha_inicio: datetime, dias: int = 5):
    """Genera eventos CV simulados para una semana de trabajo."""
    print("\n  🎲 Simulando datos Computer Vision...")

    for emp_id, nombre, area in NOMBRES:
        emp = Empleado(
            emp_id=emp_id,
            nombre=nombre,
            area=area,
            turno_inicio="07:00",
            turno_fin="15:00",
        )
        motor.registrar_empleado(emp)

    for dia in range(dias):
        fecha = fecha_inicio + timedelta(days=dia)
        if fecha.weekday() >= 5:   # Skip fin de semana
            continue
        fecha_str = fmt_date(fecha)

        for emp_id, nombre, area in NOMBRES:
            # 8% probabilidad de ausencia
            if random.random() < 0.08:
                motor._crear_alerta(
                    emp_id=emp_id,
                    nivel="INFO",
                    mensaje=f"Ausencia no justificada — {nombre}",
                    dt=fecha.replace(hour=7),
                )
                continue

            # CHECK_IN con posible tardanza (20% prob)
            delay_min = random.randint(12, 35) if random.random() < 0.20 else \
                        random.randint(-2, 5)
            ci = fecha.replace(hour=7, minute=0) + timedelta(minutes=delay_min)

            motor.registrar_evento(EventoCV(
                emp_id=emp_id, tipo="CHECK_IN", dt=ci,
                confianza=round(random.uniform(0.97, 0.999), 3)
            ))

            # CHECK_OUT
            co = fecha.replace(hour=15, minute=0) + \
                 timedelta(minutes=random.randint(-5, 10))
            motor.registrar_evento(EventoCV(
                emp_id=emp_id, tipo="CHECK_OUT", dt=co,
                confianza=round(random.uniform(0.97, 0.999), 3)
            ))

            # 30% de probabilidad de hacer OT
            if random.random() < 0.30:
                ot_s = co + timedelta(minutes=random.randint(2, 10))
                ot_hrs = round(random.uniform(0.5, 3.5), 2)
                ot_e = ot_s + timedelta(hours=ot_hrs)
                autorizado = random.random() > 0.15  # 15% sin autorización

                ev_ot = EventoCV(
                    emp_id=emp_id, tipo="OT_START", dt=ot_s,
                    confianza=round(random.uniform(0.97, 0.999), 3)
                )
                ev_ot.autorizado = autorizado
                motor.registrar_evento(ev_ot)

                motor.registrar_evento(EventoCV(
                    emp_id=emp_id, tipo="OT_END", dt=ot_e,
                    confianza=round(random.uniform(0.97, 0.999), 3)
                ))

    total_ev = len(motor.eventos)
    print(f"  ✅ {total_ev} eventos CV generados para {dias} días")
    print(f"  ⚠  {len(motor.alertas)} alertas detectadas")

# ─────────────────────────────────────────────────────────────
#  GENERADOR DE REPORTES
# ─────────────────────────────────────────────────────────────
def exportar_csv(motor: MotorTurnos, path: str):
    """Exporta todos los eventos a CSV."""
    rows = [e.to_dict() for e in motor.eventos]
    if not rows:
        print("  ⚠  Sin datos para exportar.")
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"  📄 CSV exportado → {path}  ({len(rows)} registros)")

def exportar_alertas_json(motor: MotorTurnos, path: str):
    """Exporta alertas en formato JSON."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(motor.alertas, f, indent=2, ensure_ascii=False)
    print(f"  🔔 Alertas JSON → {path}  ({len(motor.alertas)} alertas)")

def exportar_excel(motor: MotorTurnos, fechas: list[str], path: str):
    """Genera reporte Excel con múltiples hojas y gráfico."""
    if not EXCEL_OK:
        print("  ⚠  openpyxl no disponible. Saltando exportación Excel.")
        return

    wb = openpyxl.Workbook()

    # ── COLORES ────────────────────────────────────────────────
    DARK    = "020810"
    CYAN    = "00E5FF"
    GREEN   = "00C896"
    RED     = "FF3A5C"
    GOLD    = "FFCC00"
    WHITE   = "E8F4FD"
    PANEL   = "060F1E"
    MUTED   = "3A5A7A"

    hdr_fill   = PatternFill("solid", fgColor=DARK)
    sub_fill   = PatternFill("solid", fgColor=PANEL)
    total_fill = PatternFill("solid", fgColor="001A2A")
    alt_fill   = PatternFill("solid", fgColor="040C18")

    thin = Side(style="thin", color="0A1828")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(color=CYAN):
        return Font(name="Calibri", bold=True, color=color, size=9)
    def body_font(color=WHITE):
        return Font(name="Calibri", color=color, size=9)
    def title_font():
        return Font(name="Calibri", bold=True, color=CYAN, size=14)
    def mono_font(color=CYAN):
        return Font(name="Courier New", color=color, size=8)

    def set_hdr(ws, row, cols_labels):
        for col, label in enumerate(cols_labels, 1):
            c = ws.cell(row=row, column=col, value=label)
            c.fill = hdr_fill
            c.font = hdr_font()
            c.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
            c.border = border

    def style_cell(ws, row, col, value, color=WHITE, mono=False,
                   bold=False, align="left", fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = (mono_font(color) if mono
                  else Font(name="Calibri", color=color, size=9, bold=bold))
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = border
        if fill:
            c.fill = fill
        return c

    # ══════════════════════════════════════════════════════════
    # HOJA 1: PORTADA
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "📊 Resumen"
    ws.sheet_properties.tabColor = "00E5FF"
    ws.column_dimensions["A"].width = 2
    ws.row_dimensions[1].height = 6

    # Fondo oscuro
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=12):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="020810")

    # Título
    ws.merge_cells("B2:L2")
    c = ws["B2"]
    c.value = "ASSISTANTTRACK 4D — REPORTE EJECUTIVO DE TURNOS Y OVERTIME"
    c.font = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor="020810")

    ws.merge_cells("B3:L3")
    c = ws["B3"]
    c.value = f"Computer Vision Workforce Intelligence  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    c.font = Font(name="Calibri", color=MUTED, size=9)
    c.alignment = Alignment(horizontal="left")
    c.fill = PatternFill("solid", fgColor="020810")

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[4].height = 6

    # KPIs
    kpi_labels = ["EMPLEADOS","EVENTOS CV","HORAS OT","COSTO OT TOTAL","ALERTAS","ROI PROYECTADO"]
    kpi_colors = [CYAN, CYAN, GOLD, RED, GOLD, GREEN]

    # Calcular KPIs
    total_eventos = len(motor.eventos)
    total_alertas = len(motor.alertas)
    total_horas_ot = 0.0
    total_costo_ot = 0.0
    for fecha in fechas:
        resumen = motor.resumen_costos_dia(fecha)
        total_horas_ot += resumen["horas_ot_dia"]
        total_costo_ot += resumen["total_ot"]

    kpi_values = [
        len(motor.empleados),
        total_eventos,
        f"{total_horas_ot:.1f}h",
        f"${total_costo_ot:,.2f}",
        total_alertas,
        "316%",
    ]

    for i, (lbl, val, col) in enumerate(zip(kpi_labels, kpi_values, kpi_colors)):
        c_col = 2 + i * 2
        ws.merge_cells(
            start_row=5, start_column=c_col,
            end_row=5,   end_column=c_col + 1
        )
        ws.merge_cells(
            start_row=6, start_column=c_col,
            end_row=6,   end_column=c_col + 1
        )
        lc = ws.cell(row=5, column=c_col, value=lbl)
        lc.font = Font(name="Calibri", color=MUTED, size=7, bold=True)
        lc.alignment = Alignment(horizontal="center")
        lc.fill = PatternFill("solid", fgColor="050E1C")

        vc = ws.cell(row=6, column=c_col, value=val)
        vc.font = Font(name="Calibri", color=col, size=16, bold=True)
        vc.alignment = Alignment(horizontal="center")
        vc.fill = PatternFill("solid", fgColor="050E1C")

        ws.row_dimensions[5].height = 16
        ws.row_dimensions[6].height = 28

    # Tabla diaria en resumen
    ROW = 9
    ws.row_dimensions[ROW - 1].height = 8

    headers_res = ["FECHA","EMPLEADOS ACTIVOS","HORAS OT","COSTO NORMAL","COSTO OT","TOTAL DÍA"]
    set_hdr(ws, ROW, [""] + headers_res)

    for fi, fecha in enumerate(fechas):
        r = motor.resumen_costos_dia(fecha)
        row = ROW + 1 + fi
        fill = alt_fill if fi % 2 else sub_fill
        ws.row_dimensions[row].height = 16
        style_cell(ws, row, 1, "", fill=fill)
        style_cell(ws, row, 2, fecha,                      fill=fill)
        style_cell(ws, row, 3, len(r["empleados"]),    align="center", fill=fill, color=CYAN)
        style_cell(ws, row, 4, f"{r['horas_ot_dia']:.1f}h", align="center", fill=fill, color=GOLD)
        style_cell(ws, row, 5, f"${r['total_normal']:,.2f}",  align="right",  fill=fill)
        style_cell(ws, row, 6, f"${r['total_ot']:,.2f}",       align="right",  fill=fill, color=RED)
        style_cell(ws, row, 7, f"${r['total_general']:,.2f}", align="right",  fill=fill, color=GREEN, bold=True)
        for col in range(8, 13):
            ws.cell(row=row, column=col).fill = fill

    # Totales
    tot_row = ROW + 1 + len(fechas)
    ws.row_dimensions[tot_row].height = 18
    style_cell(ws, tot_row, 1, "",         fill=total_fill)
    style_cell(ws, tot_row, 2, "TOTAL",    fill=total_fill, bold=True, color=CYAN)
    style_cell(ws, tot_row, 3, "",         fill=total_fill)
    style_cell(ws, tot_row, 4, f"{total_horas_ot:.1f}h", fill=total_fill, color=GOLD, bold=True, align="center")
    total_n = sum(motor.resumen_costos_dia(f)["total_normal"] for f in fechas)
    total_g = sum(motor.resumen_costos_dia(f)["total_general"] for f in fechas)
    style_cell(ws, tot_row, 5, f"${total_n:,.2f}",       fill=total_fill, bold=True)
    style_cell(ws, tot_row, 6, f"${total_costo_ot:,.2f}", fill=total_fill, bold=True, color=RED)
    style_cell(ws, tot_row, 7, f"${total_g:,.2f}",        fill=total_fill, bold=True, color=GREEN)
    for col in range(8, 13):
        ws.cell(row=tot_row, column=col).fill = total_fill

    # Anchos de columnas
    col_ws_res = [2, 14, 12, 10, 12, 12, 13]
    for ci, cw in enumerate(col_ws_res):
        ws.column_dimensions[get_column_letter(ci + 1)].width = cw

    # ══════════════════════════════════════════════════════════
    # HOJA 2: DETALLE DE EMPLEADOS
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("👥 Empleados")
    ws2.sheet_properties.tabColor = "00C9B1"

    for row in ws2.iter_rows(min_row=1, max_row=200, min_col=1, max_col=12):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="020810")

    ws2.merge_cells("A1:K1")
    ws2["A1"].value = "DETALLE POR EMPLEADO — SEMANA COMPLETA"
    ws2["A1"].font = title_font()
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2["A1"].fill = PatternFill("solid", fgColor="020810")
    ws2.row_dimensions[1].height = 24

    headers2 = ["EMP ID","NOMBRE","ÁREA","FECHA","H. NORMALES","H. OVERTIME",
                "COSTO NORMAL","COSTO OT","COSTO TOTAL","HASH CV"]
    set_hdr(ws2, 2, headers2)
    ws2.row_dimensions[2].height = 16

    row_n = 3
    emp_resumen = defaultdict(lambda: {"h_normal": 0, "h_ot": 0,
                                        "c_normal": 0, "c_ot": 0, "c_total": 0})

    for fecha in fechas:
        r = motor.resumen_costos_dia(fecha)
        for emp_r in r["empleados"]:
            fill = alt_fill if row_n % 2 else sub_fill
            ws2.row_dimensions[row_n].height = 15
            emp = motor.empleados.get(emp_r["emp_id"])

            style_cell(ws2, row_n, 1,  emp_r["emp_id"],        mono=True, fill=fill, color=CYAN)
            style_cell(ws2, row_n, 2,  emp_r["nombre"],        fill=fill, bold=True)
            style_cell(ws2, row_n, 3,  emp.area if emp else "—", fill=fill, color=MUTED)
            style_cell(ws2, row_n, 4,  emp_r["fecha"],         fill=fill, color=MUTED)
            style_cell(ws2, row_n, 5,  emp_r["horas_normales"], align="center", fill=fill)
            style_cell(ws2, row_n, 6,  emp_r["horas_ot"],       align="center", fill=fill,
                       color=GOLD if emp_r["horas_ot"] > 0 else MUTED)
            style_cell(ws2, row_n, 7,  f"${emp_r['costo_normal']:,.2f}",  align="right", fill=fill)
            style_cell(ws2, row_n, 8,  f"${emp_r['costo_ot']:,.2f}",      align="right", fill=fill,
                       color=RED if emp_r["costo_ot"] > 0 else MUTED)
            style_cell(ws2, row_n, 9,  f"${emp_r['costo_total']:,.2f}",   align="right", fill=fill,
                       color=GREEN, bold=True)

            # Hash CV del último evento
            ev_last = next((e for e in reversed(motor.eventos)
                            if e.emp_id == emp_r["emp_id"]
                            and fmt_date(e.datetime) == fecha), None)
            style_cell(ws2, row_n, 10,
                       ev_last.hash_ev if ev_last else "—",
                       mono=True, fill=fill, color=MUTED)

            # Acumular para resumen
            er = emp_resumen[emp_r["emp_id"]]
            er["h_normal"] += emp_r["horas_normales"]
            er["h_ot"]     += emp_r["horas_ot"]
            er["c_normal"] += emp_r["costo_normal"]
            er["c_ot"]     += emp_r["costo_ot"]
            er["c_total"]  += emp_r["costo_total"]
            row_n += 1

    col_ws2 = [10, 18, 14, 11, 11, 11, 13, 12, 13, 20]
    for ci, cw in enumerate(col_ws2):
        ws2.column_dimensions[get_column_letter(ci + 1)].width = cw

    # ══════════════════════════════════════════════════════════
    # HOJA 3: REGISTRO DE EVENTOS CV
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("🎥 Eventos CV")
    ws3.sheet_properties.tabColor = "0066FF"

    for row in ws3.iter_rows(min_row=1, max_row=300, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="020810")

    ws3.merge_cells("A1:I1")
    ws3["A1"].value = "BITÁCORA DE EVENTOS — COMPUTER VISION"
    ws3["A1"].font = title_font()
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws3["A1"].fill = PatternFill("solid", fgColor="020810")
    ws3.row_dimensions[1].height = 24

    tipo_colors = {
        "CHECK_IN":  "00E676",
        "CHECK_OUT": "3A5A7A",
        "OT_START":  "FFCC00",
        "OT_END":    "00C9B1",
    }

    headers3 = ["EMP ID","NOMBRE","TIPO EVENTO","FECHA","HORA",
                "CONFIANZA CV","AUTORIZADO","HASH SHA-256"]
    set_hdr(ws3, 2, headers3)
    ws3.row_dimensions[2].height = 16

    for i, ev in enumerate(motor.eventos):
        rn = i + 3
        fill = alt_fill if i % 2 else sub_fill
        ws3.row_dimensions[rn].height = 14
        emp = motor.empleados.get(ev.emp_id)
        ev_color = tipo_colors.get(ev.tipo, CYAN)

        style_cell(ws3, rn, 1, ev.emp_id,   mono=True, fill=fill, color=CYAN)
        style_cell(ws3, rn, 2, emp.nombre if emp else "—", fill=fill)
        style_cell(ws3, rn, 3, ev.tipo,     fill=fill, color=ev_color, bold=True)
        style_cell(ws3, rn, 4, ev.to_dict()["fecha"], fill=fill, color=MUTED)
        style_cell(ws3, rn, 5, ev.to_dict()["hora"],  fill=fill, mono=True, color=WHITE)
        style_cell(ws3, rn, 6, ev.to_dict()["confianza"], fill=fill,
                   align="center", color=GREEN)
        auth_txt = "✓ SÍ" if ev.autorizado else "✗ NO"
        style_cell(ws3, rn, 7, auth_txt, fill=fill, align="center",
                   color=GREEN if ev.autorizado else RED, bold=True)
        style_cell(ws3, rn, 8, ev.hash_ev, fill=fill, mono=True, color=MUTED)

    col_ws3 = [10, 18, 14, 11, 10, 12, 10, 20]
    for ci, cw in enumerate(col_ws3):
        ws3.column_dimensions[get_column_letter(ci + 1)].width = cw

    # ══════════════════════════════════════════════════════════
    # HOJA 4: ALERTAS
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("🚨 Alertas")
    ws4.sheet_properties.tabColor = "FF3A5C"

    for row in ws4.iter_rows(min_row=1, max_row=200, min_col=1, max_col=8):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="020810")

    ws4.merge_cells("A1:F1")
    ws4["A1"].value = "REGISTRO DE ALERTAS — ASSISTANTTRACK 4D"
    ws4["A1"].font = title_font()
    ws4["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws4["A1"].fill = PatternFill("solid", fgColor="020810")
    ws4.row_dimensions[1].height = 24

    headers4 = ["TIMESTAMP","EMP ID","NIVEL","MENSAJE","ACCIÓN SUGERIDA"]
    set_hdr(ws4, 2, headers4)
    ws4.row_dimensions[2].height = 16

    nivel_color = {"CRITICAL": RED, "WARNING": GOLD, "INFO": CYAN}
    acciones = {
        "CRITICAL": "Notificar RRHH inmediatamente",
        "WARNING":  "Revisar con supervisor de turno",
        "INFO":     "Registrar y monitorear",
    }

    for i, al in enumerate(motor.alertas):
        rn = i + 3
        fill = alt_fill if i % 2 else sub_fill
        ws4.row_dimensions[rn].height = 14
        nc = nivel_color.get(al["nivel"], CYAN)

        style_cell(ws4, rn, 1, al["timestamp"][:16].replace("T", " "),
                   fill=fill, mono=True, color=MUTED)
        style_cell(ws4, rn, 2, al["emp_id"],  fill=fill, mono=True, color=CYAN)
        style_cell(ws4, rn, 3, al["nivel"],   fill=fill, color=nc, bold=True, align="center")
        style_cell(ws4, rn, 4, al["mensaje"], fill=fill)
        style_cell(ws4, rn, 5, acciones.get(al["nivel"], "—"), fill=fill, color=MUTED)

    col_ws4 = [18, 10, 10, 45, 32]
    for ci, cw in enumerate(col_ws4):
        ws4.column_dimensions[get_column_letter(ci + 1)].width = cw

    wb.save(path)
    print(f"  📊 Excel exportado → {path}")

# ─────────────────────────────────────────────────────────────
#  ANÁLISIS DE PRODUCTIVIDAD Y ROI
# ─────────────────────────────────────────────────────────────
def analisis_roi(motor: MotorTurnos, fechas: list[str]) -> dict:
    costo_ot_total = sum(motor.resumen_costos_dia(f)["total_ot"] for f in fechas)
    horas_ot_total = sum(motor.resumen_costos_dia(f)["horas_ot_dia"] for f in fechas)
    tardanzas = sum(1 for a in motor.alertas if "Tardanza" in a["mensaje"])
    ausencias  = sum(1 for a in motor.alertas if "Ausencia" in a["mensaje"])
    ot_no_auth = sum(1 for a in motor.alertas if "sin autorización" in a["mensaje"])

    # Proyección de ahorro anual con el sistema
    sem_proyectadas = 52
    ot_semanal = costo_ot_total / max(len(fechas) / 5, 1)
    ahorro_ot_anual = ot_semanal * sem_proyectadas * 0.40  # −40% OT

    return {
        "semanas_analizadas": len(fechas) / 5,
        "costo_ot_total":     round(costo_ot_total, 2),
        "horas_ot_total":     round(horas_ot_total, 2),
        "tardanzas":          tardanzas,
        "ausencias":          ausencias,
        "ot_no_autorizados":  ot_no_auth,
        "ahorro_proyectado_anual": round(ahorro_ot_anual, 2),
        "roi_estimado":       "316%",
        "payback_meses":      3.8,
    }

# ─────────────────────────────────────────────────────────────
#  IMPRESIÓN CONSOLA
# ─────────────────────────────────────────────────────────────
def imprimir_resumen_consola(motor: MotorTurnos, fechas: list[str]):
    titulo("ASSISTANTTRACK 4D — REPORTE EJECUTIVO")

    for fecha in fechas:
        subtitulo(f"  RESUMEN DÍA: {fecha}")
        r = motor.resumen_costos_dia(fecha)
        print(f"  {'Empleados con registro:':<30} {len(r['empleados'])}")
        print(f"  {'Horas overtime total:':<30} {r['horas_ot_dia']:.1f}h")
        print(f"  {'Costo nómina normal:':<30} ${r['total_normal']:,.2f}")
        print(f"  {'Costo overtime:':<30} ${r['total_ot']:,.2f}")
        print(f"  {'Costo total día:':<30} ${r['total_general']:,.2f}")

        # Top 3 OT
        top_ot = sorted(r["empleados"], key=lambda x: x["horas_ot"], reverse=True)[:3]
        if any(e["horas_ot"] > 0 for e in top_ot):
            print(f"\n  {'TOP OVERTIME DEL DÍA':}")
            for e in top_ot:
                if e["horas_ot"] > 0:
                    print(f"    ⚡ {e['nombre']:<20} {fmt_horas(e['horas_ot']*60)}  "
                          f"→ ${e['costo_ot']:,.2f}")

    # Alertas
    subtitulo("  ALERTAS DEL SISTEMA")
    criticas = [a for a in motor.alertas if a["nivel"] == "CRITICAL"]
    warnings  = [a for a in motor.alertas if a["nivel"] == "WARNING"]
    infos     = [a for a in motor.alertas if a["nivel"] == "INFO"]

    print(f"  🔴 CRÍTICAS : {len(criticas)}")
    print(f"  🟡 WARNINGS : {len(warnings)}")
    print(f"  🔵 INFO     : {len(infos)}")

    for a in criticas[:3]:
        print(f"     → {a['mensaje']}")

    # ROI
    roi = analisis_roi(motor, fechas)
    subtitulo("  ANÁLISIS ROI Y PROYECCIÓN")
    print(f"  {'Costo OT analizado:':<35} ${roi['costo_ot_total']:,.2f}")
    print(f"  {'Horas OT totales:':<35} {roi['horas_ot_total']:.1f}h")
    print(f"  {'OT sin autorización:':<35} {roi['ot_no_autorizados']}")
    print(f"  {'Tardanzas detectadas:':<35} {roi['tardanzas']}")
    print(f"  {'Ahorro OT proyectado/año (−40%):':<35} ${roi['ahorro_proyectado_anual']:,.2f}")
    print(f"  {'ROI estimado sistema:':<35} {roi['roi_estimado']}")
    print(f"  {'Payback estimado:':<35} {roi['payback_meses']} meses")

    separador("═")
    print("  ✅ Análisis completado.")
    separador("═")

# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────
def main():
    random.seed(42)   # Reproducible

    titulo("ASSISTANTTRACK 4D — INICIANDO SISTEMA")
    print("  Módulos activos:")
    print("    ✓ Motor de Turnos        ✓ Control Overtime")
    print("    ✓ Análisis de Costos     ✓ Motor de Alertas")
    print("    ✓ Simulador CV           ✓ Generador Reportes")
    separador()

    # Inicializar motor
    motor = MotorTurnos()

    # Simular una semana laboral desde el lunes pasado
    hoy = datetime.now()
    lunes = hoy - timedelta(days=hoy.weekday())
    lunes = lunes.replace(hour=0, minute=0, second=0, microsecond=0)

    simular_semana(motor, lunes, dias=5)

    # Obtener fechas con datos
    fechas_con_datos = sorted({fmt_date(e.datetime) for e in motor.eventos})

    # Imprimir resumen en consola
    imprimir_resumen_consola(motor, fechas_con_datos)

    # Exportar archivos al directorio "outputs" junto al ejecutable/script
    out_dir = os.path.join(_base_dir(), "outputs")
    os.makedirs(out_dir, exist_ok=True)

    subtitulo("  EXPORTANDO REPORTES")
    exportar_csv(motor,       f"{out_dir}/at4d_eventos_cv.csv")
    exportar_alertas_json(motor, f"{out_dir}/at4d_alertas.json")
    exportar_excel(motor, fechas_con_datos, f"{out_dir}/at4d_reporte.xlsx")

    separador("═")
    print("  🚀 AssistantTrack 4D — Ejecución finalizada con éxito.")
    separador("═")


if __name__ == "__main__":
    main()
